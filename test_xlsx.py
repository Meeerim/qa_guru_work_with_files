import os
from openpyxl import load_workbook
from os_path.os_path_scripts import resources

# TODO оформить в тест, добавить ассерты и использовать универсальный путь
xlsx_filepath = os.path.join(resources, 'file_example_XLSX_50.xlsx')


def test_xlsx_document():
    workbook = load_workbook(xlsx_filepath)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert sheet.min_column >= 1
    assert sheet.max_row == 51
    assert sheet.cell(row=5, column=3).value == "Hanner"
    assert sheet.cell(row=44, column=8).value == 3259
